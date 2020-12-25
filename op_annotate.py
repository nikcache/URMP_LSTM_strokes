#Nikesh 
#OpenPose relative movement to annotations

#Importing modules
import openpyxl as xl
import xlrd
import os
from collections import deque
import numpy as np

def annotate(class_lst, in_path, out_path, joint, a_dir, threshold, maxAvg, rollAvg):

    for i in os.listdir(in_path):
        if i.endswith('.xlsx'):
            in_file = in_path + i
            break
    
    wb = xl.load_workbook(in_file)
    ws = wb.active

    accel_list = [0]

    if a_dir == 'x':
        r_column = joint * 2 + 1
    elif a_dir == 'y':
        r_column = joint * 2 + 2
    else:
        print('Invalid direction value')
        return None

    for j in range(ws.max_row):
        accel = ws.cell(row = j + 1, column = r_column).value
        accel_list.append(accel)

    accel_label = val2label(class_lst, accel_list, threshold, maxAvg)
    # avg_accel_label = rollingAvg(accel_label, rollAvg, class_lst)

    try:
        vid_name = i[i.find('Vid'):i.rfind('.')]
        os.mkdir(out_path + vid_name)
    except:
        pass

    wb.close()
    
    outTxt = open(out_path + vid_name + '/' + vid_name + '_strokes.txt', 'w')
    outTxt.write(str(accel_label))
    outTxt.close()

    # print('Done with ' + str(i))

def rollingAvg(lst, rollAvg, class_lst):

    ret_lst = []
    Q = deque(maxlen=rollAvg)

    for i in lst:
        if i == class_lst[0]:
            k = 0
        elif i == class_lst[1]:
            k = 1
        else:
            k = 2
    
        Q.append(k)
        results = int(round(np.array(Q).mean(axis=0)))
        ret_lst.append(class_lst[results])

    return ret_lst

def val2label(class_lst, lst, threshold, maxAvg):

    ret_lst = []
    Q = deque(maxlen=maxAvg)

    for i in lst:
        Q.append(i)
        results = np.array(Q)
        j = np.argmax(results)
        results = results[j]
        # print(results)
        if results < -threshold or results > threshold or results == 0:
            if results < 0:
                #print('Appended down', results)
                ret_lst.append(class_lst[0])
            elif results > 0:
                ret_lst.append(class_lst[1])
            else:
                ret_lst.append(class_lst[2])  
        else:
            ret_lst.append(class_lst[2])
            
    return ret_lst 

def main(inst):

    #Changeable values
    class_lst = ['up_stroke', 'down_stroke', 'no_stroke']
    root_dir = 'vn_accel_J'
    outPath = 'vn_stroke_set'
    joint = 4
    a_dir = 'y'
    threshold = 2
    maxAvg = 2
    rollAvg = 0

    #Do not change unless necessary
    try:
        os.mkdir(outPath)
    except:
        pass
    
    outPath = outPath + '/'
    for vid in os.listdir(root_dir):
        inPath = root_dir + '/' + vid + '/'
        annotate(class_lst, inPath, outPath, joint, a_dir, threshold, maxAvg, rollAvg)
        print('Done with ' + vid)
    
    print('Done with ' + inst)

if __name__ == "__main__":
    main('Violin')