#Nikesh Ghimire
#Summer Research 2020
#Optical Flow Dataset Maker

#Importing Modules
import cv2
import numpy as np
import json
import openpyxl as xl
import os 
import csv
import pandas
import shutil
from random import randrange


#Extracts frames from  video in name, with instName as name and location of video
#Places frames in either playing or n_playing folder depending on silences.json
def frameOut(path, outPathRaw, anno_path, frame_lim):             

    try:

        for i in os.listdir(path):
            if i.endswith('.mp4'):# or i.endswith('.MOV'):
                vid = path + i
                break
        
        fourcc = cv2.VideoWriter_fourcc(*'DIVX')

        cap = cv2.VideoCapture(vid)
        ret, frame1 = cap.read()
        prvs = cv2.cvtColor(frame1,cv2.COLOR_BGR2GRAY)
        hsv = np.zeros_like(frame1)
        hsv[...,1] = 255

        counts = {'no_stroke': 0, 'up_stroke': 0, 'down_stroke': 0}
        frame_num = 0
        currCount = 0

        label_lst = getLabels(anno_path)
        is_begin = True

        # output_file = outPathRaw + label_lst[0] + '/' + i[:i.rfind('.')] + '_' + str(vid_count) + '.avi'
        # out = cv2.VideoWriter(output_file, fourcc, 30, (w, h), True)
    
    except Exception as e:

        print(e)
        pass

    try:
        while(1):
            ret, frame2 = cap.read()
            processed = frame2
            frame_label = label_lst[frame_num]
            old_label = label_lst[frame_num-1]

            if frame_label != old_label:
                is_begin = True

            if is_begin:# and currCount > 5:
                output_file = outPathRaw + frame_label + '/' + i[:i.rfind('.')] + '_' + str(counts[frame_label]) + '.avi'
                counts[frame_label]
                if currCount > frame_lim:
                    counts[old_label] += 1
                h, w, _ = processed.shape
                out = cv2.VideoWriter(output_file, fourcc, 30, (w, h), True)
                is_begin = False
                currCount = 0

            next = cv2.cvtColor(frame2,cv2.COLOR_BGR2GRAY)

            if frame2 is None:
                break

            flow = cv2.calcOpticalFlowFarneback(prvs,next, None, 0.5, 3, 7, 3, 5, 1.2, 0)

            threshold = 2
            # flow[np.abs(flow) < threshold] = 0

            # mag, ang = cv2.cartToPolar(flow[...,0], flow[...,1])
            # hsv[...,0] = ang*180/np.pi/2
            # hsv[...,2] = cv2.normalize(mag,None,0,255,cv2.NORM_MINMAX)
            # rgb = cv2.cvtColor(hsv,cv2.COLOR_HSV2BGR)

            # dense_flow = cv2.addWeighted(frame, 1,rgb, 2, 0)

            out.write(frame2)
            cv2.imshow('frame2',frame2)
            k = cv2.waitKey(1) & 0xff
            if k == 27:
                break
            elif k == ord('s'):
                cv2.imwrite('opticalfb.png',frame2)

            #Check playing or not playing from JSON


            # cv2.imwrite(outPath + frame_label +'/op_' + dir_name + '_' + str(frame_num) + '_' + frame_label[:2] + '.jpg',rgb)
            # cv2.imwrite(outPathRaw + frame_label +'/raw_' + dir_name + '_' + str(frame_num) + '_' + frame_label[:2] + '.jpg',frame2)
            frame_num += 1
            currCount += 1
            prvs = next

        cap.release()
        cv2.destroyAllWindows()
    
    except Exception as e:
        print(e)
        pass

#HELPER FUNCTION
#Gets pregenerated labels stored in txt
def getLabels(anno_path):

    ext = '.txt'
    for i in os.listdir(anno_path):
        if i.endswith(ext):
            anno_file = i
            break
    
    anno_file = open(anno_path + anno_file)
    anno_lst = eval(anno_file.read())

    return anno_lst

#Compiles silence intervals from the json files
def compileInterval(path):

    for i in os.listdir(path):
        if i.endswith('.JSON'):
            break

    with open(path + i, 'r', encoding='utf-8-sig') as inFile:
        data = json.load(inFile)

    s_list = [[round(eval(k),2), round(v,2)] for k, v in data['silences'].items()] 
    
    return s_list

#HELPER FUNCTION
#Returns boolean value of whether the current frame is playing or n_playing
def checkPlay(int_list, currFrame):

    currTime = currFrame/30

    for i in int_list:
        if currTime > i[0] and currTime < i[1]:
            return False
        else:
            pass
    
    return True

#Compiles a CSV for the folder path and instrument instName
def compileCSV(path, instName):

    instPath = path + '/' + instName + '/'

    up_lst = os.listdir(instPath + 'up_stroke')
    down_lst = os.listdir(instPath + 'down_stroke')
    no_lst = os.listdir(instPath + 'no_stroke')

    wb = xl.Workbook()
    ws = wb.active

    ws.cell(row = 1, column = 1).value = 'name'
    ws.cell(row = 1, column = 2).value = 'label'

    for i in no_lst:
        row_num = ws.max_row  + 1
        ws.cell(row = row_num, column = 1).value = 'no_stroke/' + i
        ws.cell(row = row_num, column = 2).value = 'no_stroke' 

    for i in down_lst:
        row_num = ws.max_row  + 1
        ws.cell(row = row_num, column = 1).value = 'down_stroke/' + i
        ws.cell(row = row_num, column = 2).value = 'down_stroke'

    for i in up_lst:
        row_num = ws.max_row  + 1
        ws.cell(row = row_num, column = 1).value = 'up_stroke/' + i
        ws.cell(row = row_num, column = 2).value = 'up_stroke'

    wb.save(instPath + '/cleaned_pre.xlsx')

    data_xls = pandas.read_excel(instPath + '/cleaned_pre.xlsx', 'Sheet', index_col=None)
    data_xls.to_csv(instPath + '/cleaned.csv', encoding='utf-8', index=False)

    for i in os.listdir(instPath):
        if i.endswith('.xlsx'):
            os.remove(instPath + '/' + i)

def move(rootDir, numSamples, class_lst):

    try:
        os.mkdir(rootDir + 'unused')
        for i in class_lst:
            os.mkdir(rootDir + 'unused/' + i)
    except:
        print('Cannot create directories or they already exist')
#
    up_lst = os.listdir(rootDir + 'up_stroke')
    down_lst = os.listdir(rootDir + 'down_stroke')
    no_lst = os.listdir(rootDir + 'no_stroke')
    moved = []

    for i in range(numSamples):
        u, d, n = randrange(len(up_lst)), randrange(len(down_lst)), randrange(len(no_lst))
        while up_lst[u] in moved:
            u = randrange(len(up_lst))
        while down_lst[d] in moved:
            d = randrange(len(down_lst))
        while no_lst[n] in moved:
            n = randrange(len(no_lst))
        os.rename(rootDir + class_lst[0] + '/' + up_lst[u], rootDir + 'unused/' + class_lst[0] + '/' + up_lst[u])
        os.rename(rootDir + class_lst[1] + '/' + down_lst[d], rootDir + 'unused/' + class_lst[1] + '/' + down_lst[d])
        os.rename(rootDir + class_lst[2] + '/' + no_lst[n], rootDir + 'unused/' + class_lst[2] + '/' + no_lst[n])
        moved.append(up_lst[u])
        moved.append(down_lst[d])
        moved.append(no_lst[n])

    print('Successfully moved ' + str(numSamples) + ' files to another folder')
#Runs frameOut function on all videos for a specific instrument. This gives use p or np frames in their respective folders 
def extract_flow_inst_LSTM(frame_lim, rootDir, vidRange, out_path_raw, anno_loc, class_lst):

    try:
        # os.mkdir(out_path)
        os.mkdir(out_path_raw)
        for i in class_lst:
            # os.mkdir(out_path + i)
            os.mkdir(out_path_raw + i)
    except:
        pass  

    for i in os.listdir(rootDir)[vidRange[0]:vidRange[1]]:
        n_path = rootDir + i + '/'
        anno_path = anno_loc + i + '/'
        frameOut(n_path, out_path_raw, anno_path, frame_lim)
        for i in class_lst:
            vid_list = os.listdir(out_path_raw + i)
            os.remove(out_path_raw + i + '/' + vid_list[-1])
    # compileCSV('URMP/data/', inst)

def main():

    inst = 'Violin_t=0'
    inst_dir = 'vn_dataset/'
    train_dir = ['LSTM_N/raw/train/'] 
    val_dir = ['LSTM_N/raw/validate/']
    anno_loc = 'vn_stroke_set/'
    class_lst = ['up_stroke', 'down_stroke', 'no_stroke']
    frame_lim = 7

    train_range = [0,6]
    val_range = [6,7]

    try:
        os.mkdir('LSTM_N/raw')
    except:
        pass

    extract_flow_inst_LSTM(frame_lim, inst_dir, train_range, train_dir[0], anno_loc, class_lst)
    extract_flow_inst_LSTM(frame_lim, inst_dir, val_range, val_dir[0], anno_loc, class_lst)

if __name__ == '__main__':

    main()